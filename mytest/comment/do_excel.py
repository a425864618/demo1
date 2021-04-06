from openpyxl import load_workbook


class DoExcel(object):
    def __init__(self, fildname):
        self.fildname = fildname
        self.wb = load_workbook(fildname)

    def ReadExcel(self, sheetname, case_id=None):
        get_sheet = self.wb[sheetname]
        list1 = []
        for i in range(2, get_sheet.max_row + 1):
            dict1 = {}
            dict1['id'] = get_sheet.cell(row=i, column=1).value
            dict1['title'] = get_sheet.cell(row=i, column=2).value
            dict1['method'] = get_sheet.cell(row=i, column=3).value
            dict1['url'] = get_sheet.cell(row=i, column=4).value
            dict1['run'] = get_sheet.cell(row=i, column=5).value
            dict1['data'] = get_sheet.cell(row=i, column=6).value
            dict1['assert'] = get_sheet.cell(row=i, column=7).value
            dict1['real_result'] = get_sheet.cell(row=i, column=8).value
            dict1['is_pass'] = get_sheet.cell(row=i, column=9).value
            dict1['update_time'] = get_sheet.cell(row=i, column=10).value
            dict1['sql_form'] = get_sheet.cell(row=i, column=11).value
            dict1['is_depend'] = get_sheet.cell(row=i, column=12).value
            dict1['depend_cookies'] = get_sheet.cell(row=i, column=13).value
            dict1['depend_id'] = get_sheet.cell(row=i, column=14).value
            dict1['depend_data'] = get_sheet.cell(row=i, column=15).value




            list1.append(dict1)
        if case_id != None:
            for i in list1:
                if case_id == i["id"]:
                    return i
        else:
            return list1

    def Write(self, sheetname, row, real_result, is_pass, update_time):
        get_sheet = self.wb[sheetname]
        get_sheet.cell(row, 8).value = real_result
        get_sheet.cell(row, 9).value = is_pass
        get_sheet.cell(row, 10).value = update_time
        self.wb.save(self.fildname)


if __name__ == '__main__':
    a = DoExcel("../data/cases.xlsx")
    # Sheet
    b = a.ReadExcel('充值')
    for i in b:
        print(i)
